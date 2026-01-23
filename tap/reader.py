"""Excel/CSV文件读取模块"""

from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import load_workbook
import csv


class ExcelReader:
    """Excel文件读取器"""
    
    def __init__(self, file_path: str, frozen_zone: str = "0:5", data_zone: str = "6:25"):
        """
        初始化Excel读取器
        
        Args:
            file_path: Excel文件路径
            frozen_zone: 冻结区域列范围，数字索引，格式: start:end (如"0:5"表示第0到5列，对应A-F)
            data_zone: 数据区域列范围，数字索引，格式: start:end (如"6:25"表示第6到25列，对应G-Z)
        """
        self.file_path = Path(file_path)
        self.frozen_zone = frozen_zone
        self.data_zone = data_zone
        
        self._frozen_cols = self._parse_zone(frozen_zone)
        self._data_cols = self._parse_zone(data_zone)
    
    def _parse_zone(self, zone: str) -> Tuple[int, int]:
        """解析区域字符串，数字索引格式
        支持格式:
        - "start:end" -> (start, end) 如 "0:5"
        - "single" -> (single, single) 如 "5184" 表示只有一列
        """
        if ":" in zone:
            start, end = zone.split(":")
            return (int(start), int(end))
        else:
            # 单个数字，表示只有一列
            single = int(zone)
            return (single, single)
    
    def _col_to_index(self, col: str) -> int:
        """将列字母转换为索引（0-based）"""
        result = 0
        for char in col.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1
    
    def _index_to_col(self, index: int) -> str:
        """将索引转换为列字母"""
        result = ""
        index += 1
        while index > 0:
            index -= 1
            result = chr(ord('A') + index % 26) + result
            index //= 26
        return result
    
    def read_headers(self, sheet_index: int = 0) -> List[str]:
        """读取表头（第一行）"""
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.active
        
        headers = []
        for col in range(self._data_cols[0], self._data_cols[1] + 1):
            cell_value = ws.cell(row=1, column=col + 1).value
            headers.append(str(cell_value) if cell_value is not None else "")
        
        wb.close()
        return headers
    
    def read_frozen_headers(self, sheet_index: int = 0) -> List[str]:
        """读取冻结区域表头"""
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.active
        
        headers = []
        for col in range(self._frozen_cols[0], self._frozen_cols[1] + 1):
            cell_value = ws.cell(row=1, column=col + 1).value
            headers.append(str(cell_value) if cell_value is not None else "")
        
        wb.close()
        return headers
    
    def read_data(self, sheet_index: int = 0) -> List[Dict[str, Any]]:
        """读取数据区域数据"""
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.active
        
        data = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            row_data = {}
            for col in range(self._data_cols[0], self._data_cols[1] + 1):
                col_letter = self._index_to_col(col)
                header = ws.cell(row=1, column=col + 1).value
                if header:
                    row_data[header] = row[col].value
            data.append(row_data)
        
        wb.close()
        return data
    
    def read_frozen_data(self, sheet_index: int = 0) -> List[Dict[str, Any]]:
        """读取冻结区域数据"""
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.active
        
        data = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            row_data = {}
            for col in range(self._frozen_cols[0], self._frozen_cols[1] + 1):
                header = ws.cell(row=1, column=col + 1).value
                if header:
                    row_data[header] = row[col].value
            data.append(row_data)
        
        wb.close()
        return data
    
    def read_all(self, sheet_index: int = 0) -> Tuple[List[str], List[Dict[str, Any]], List[Dict[str, Any]]]:
        """读取所有数据"""
        frozen_headers = self.read_frozen_headers(sheet_index)
        data_headers = self.read_headers(sheet_index)
        frozen_data = self.read_frozen_data(sheet_index)
        data_rows = self.read_data(sheet_index)
        
        return frozen_headers, frozen_data, data_rows


class CSVReader:
    """CSV文件读取器"""
    
    def __init__(self, file_path: str, frozen_cols: Optional[Tuple[int, int]] = None, 
                 data_cols: Optional[Tuple[int, int]] = None):
        self.file_path = Path(file_path)
        self.frozen_cols = frozen_cols or (0, 5)  # 默认A-F
        self.data_cols = data_cols or (6, 25)     # 默认G-Z
    
    def read_headers(self) -> List[str]:
        """读取数据区域表头"""
        with open(self.file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            headers = list(reader)[0]
            # 去除BOM字符
            headers = [h.lstrip('\ufeff') if isinstance(h, str) else h for h in headers]
            return headers[self.data_cols[0]:self.data_cols[1] + 1]
    
    def read_frozen_headers(self) -> List[str]:
        """读取冻结区域表头"""
        with open(self.file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            headers = list(reader)[0]
            # 去除BOM字符
            headers = [h.lstrip('\ufeff') if isinstance(h, str) else h for h in headers]
            return headers[self.frozen_cols[0]:self.frozen_cols[1] + 1]
    
    def read_data(self) -> List[Dict[str, Any]]:
        """读取数据区域数据"""
        with open(self.file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            all_rows = list(reader)
            if not all_rows:
                return []
            
            headers = all_rows[0]
            # 去除BOM字符
            headers = [h.lstrip('\ufeff') if isinstance(h, str) else h for h in headers]
            data_headers = headers[self.data_cols[0]:self.data_cols[1] + 1]
            
            data = []
            for row in all_rows[1:]:
                if len(row) > self.data_cols[0]:
                    row_data = {}
                    for i, header in enumerate(data_headers):
                        idx = self.data_cols[0] + i
                        if idx < len(row):
                            row_data[header] = row[idx]
                    data.append(row_data)
            return data
    
    def read_frozen_data(self) -> List[Dict[str, Any]]:
        """读取冻结区域数据"""
        with open(self.file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            all_rows = list(reader)
            if not all_rows:
                return []
            
            headers = all_rows[0]
            # 去除BOM字符
            headers = [h.lstrip('\ufeff') if isinstance(h, str) else h for h in headers]
            frozen_headers = headers[self.frozen_cols[0]:self.frozen_cols[1] + 1]
            
            data = []
            for row in all_rows[1:]:
                if len(row) > self.frozen_cols[0]:
                    row_data = {}
                    for i, header in enumerate(frozen_headers):
                        idx = self.frozen_cols[0] + i
                        if idx < len(row):
                            row_data[header] = row[idx]
                    data.append(row_data)
            return data


def _parse_zone_tuple(zone: str) -> Tuple[int, int]:
    """解析区域字符串为元组
    支持格式:
    - "start:end" -> (start, end) 如 "0:5"
    - "single" -> (single, single) 如 "5184" 表示只有一列
    """
    if ":" in zone:
        start, end = zone.split(":")
        return (int(start), int(end))
    else:
        single = int(zone)
        return (single, single)


def get_reader(file_path: str, frozen_zone: str = "0:5", data_zone: str = "6:25"):
    """获取合适的文件读取器
    
    Args:
        file_path: 文件路径
        frozen_zone: 冻结区域列范围，数字索引格式 (如 "0:5")
        data_zone: 数据区域列范围，数字索引格式 (如 "6:25" 或 "5184" 表示单列)
    """
    path = Path(file_path)
    if path.suffix.lower() == '.csv':
        # 解析数字索引格式
        frozen_cols = _parse_zone_tuple(frozen_zone)
        data_cols = _parse_zone_tuple(data_zone)
        return CSVReader(file_path, frozen_cols, data_cols)
    else:
        return ExcelReader(file_path, frozen_zone, data_zone)
